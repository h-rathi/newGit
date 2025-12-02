#!/usr/bin/env python3
"""
Merged scraper script: Amazon, BestBuy, Samsung
Keeps each site's original logic (parsing, waits, mouse moves, cookie handling) intact.
This version additionally saves the run's result as a single row in an Excel file:
- Columns = keys (flattened per-site/per-index keys like "amazon_1_url")
- Row = values for this run
- On next run the script appends a new row (does not overwrite previous data).
Only code returned as requested.
"""
import asyncio
import json
import os
import random
import re
from datetime import datetime
from urllib.parse import quote_plus
from playwright.async_api import async_playwright, TimeoutError
from bs4 import BeautifulSoup
from openpyxl.utils import column_index_from_string
# new imports for Excel writing
from openpyxl import Workbook, load_workbook

# -----------------------
# Shared helpers
# -----------------------
async def human_delay(min_sec=0.5, max_sec=2.5):
    """Wait for a random time between min_sec and max_sec seconds."""
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)

async def human_delay_short():
    """Small helper to yield control briefly (kept minimal to respect original logic)."""
    await asyncio.sleep(0.1)

def sanitize_filename(s: str, maxlen: int = 200) -> str:
    """Create a filesystem-safe short filename from a string (URL)."""
    if not s:
        return "file"
    s_enc = quote_plus(s, safe="")
    s_clean = re.sub(r'[^A-Za-z0-9._-]', '_', s_enc)
    return s_clean[:maxlen]

def _to_jsonable(v):
    """Convert complex types to JSON strings for Excel storage; leave primitives as-is."""
    if isinstance(v, (dict, list, tuple)):
        return json.dumps(v, ensure_ascii=False)
    return v

def save_dict_to_excel_row(data: dict, excel_path: str = "outputs/results.xlsx"):
    """
    Save the provided dict as a single row in an Excel file.
    - Keys become column headers (first row).
    - Values become the next available row.
    - If the file exists, new keys are appended as new columns; existing column order is preserved.
    """
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        headers = list(data.keys())
        ws.append(headers)
        row = [ _to_jsonable(data.get(h)) for h in headers ]
        ws.append(row)
        wb.save(excel_path)
        print(f"✅ Results written to new Excel file: {excel_path}")
        return

    # file exists - load and append
    wb = load_workbook(excel_path)
    ws = wb.active

    # read existing headers from first row
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    existing_headers = [cell.value for cell in first_row]

    # compute headers union preserving existing order and appending new keys at the end
    new_keys = [k for k in data.keys() if k not in existing_headers]
    if new_keys:
        headers = existing_headers + new_keys
        # rewrite header row with expanded headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        headers = existing_headers

    # build row in header order
    row = []
    for h in headers:
        v = data.get(h)
        row.append(_to_jsonable(v) if v is not None else None)

    ws.append(row)
    wb.save(excel_path)

    #making changes from here
    column_refs = [
        "a","d","ar","x","e","as","y","blank column",
        "i","aw","ac","j","ax","ad","blank column",
        "n","bb","ah","o","bc","ai","blank column",
        "s","bg","am","t","bh","an"
    ]
    new_sheet_base_name="SelectedColumns"
    source_sheet_name=None
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    # new_name = new_sheet_base_name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    # tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1
    wb.save(excel_path)
    print(f"✅ Results appended to Excel file: {excel_path}")





def copy_columns_by_references(
    file_path: str,
    column_refs: list,
    source_sheet_name: str | None = None,
    new_sheet_base_name: str = "CopiedColumns"
) -> str:
    """
    Copy columns from source sheet to a new sheet using Excel column letters.
    - column_refs: list of strings, column letters like ['A','D','X','AR', ...] or 'blank column'
    - source_sheet_name: None -> first sheet is used
    Returns the name of the created sheet.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path)
    # select source sheet
    if source_sheet_name:
        if source_sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{source_sheet_name}' not found in workbook.")
        src = wb[source_sheet_name]
    else:
        src = wb[wb.sheetnames[0]]

    # create unique new sheet name
    new_name = "converted"
    if new_name in wb.sheetnames:
        del wb["converted"]
    # i = 1
    # while new_name in wb.sheetnames:
    #     new_name = f"{new_sheet_base_name}_{i}"
    #     i += 1
    tgt = wb.create_sheet(title=new_name)

    max_row = src.max_row if src.max_row is not None else 0

    # target column pointer (1-indexed for openpyxl)
    tgt_col_idx = 1

    for token in column_refs:
        is_blank = token is None or (isinstance(token, str) and token.strip().lower() == "blank column")
        if is_blank:
            # leave a blank column (i.e., do nothing but advance tgt_col_idx)
            tgt_col_idx += 1
            continue

        # try to interpret token as Excel column letters
        col_letters = str(token).strip()
        try:
            src_col_idx = column_index_from_string(col_letters.upper())
        except Exception:
            # invalid column reference — create an empty column instead
            for r in range(1, max_row + 1):
                tgt.cell(row=r, column=tgt_col_idx, value=None)
            tgt_col_idx += 1
            continue

        # Copy values from source column to target column
        for r in range(1, max_row + 1):
            src_cell = src.cell(row=r, column=src_col_idx)
            # copy value only (not style/formula). If formula needed, assign src_cell.value (it will copy the formula text)
            tgt.cell(row=r, column=tgt_col_idx, value=src_cell.value)
        tgt_col_idx += 1

    # Save workbook (overwrites existing file)
    wb.save(file_path)
    return new_name


# -----------------------
# AMAZON-specific logic
# -----------------------
async def save_amazon_htmls(
    urls,
    output_dir="outputs",
    cookies_file="amazon_cookies.json",
    headless=True,
):
    """Loop over the list of URLs, save each HTML to a unique file, and update cookies once."""
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless, slow_mo=100)

        # Load existing cookies/session state if available
        if os.path.exists(cookies_file):
            print("🍪 Loading existing cookies/session...")
            context = await browser.new_context(storage_state=cookies_file)
        else:
            print("🆕 No cookies found, creating a new session...")
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 768},
            )

        try:
            results = []
            for idx, url in enumerate(urls, start=1):
                try:
                    safe_name = sanitize_filename(url)[:120]
                    output_file = os.path.join(output_dir, f"amazon_{idx}_{safe_name}.html")

                    page = await context.new_page()
                    print(f"\n[Amazon {idx}/{len(urls)}] Navigating to {url} ...")
                    await page.goto(url, wait_until="load")

                    # Wait randomly for page content to settle
                    await human_delay(3, 6)

                    # 🖱️ Simulate random human-like mouse movement
                    for _ in range(3):
                        x = random.randint(200, 800)
                        y = random.randint(200, 600)
                        await page.mouse.move(x, y, steps=random.randint(5, 15))
                        await human_delay(0.3, 1.5)

                    # 🖱️ Random scrolling
                    for _ in range(2):
                        scroll_y = random.randint(400, 1000)
                        await page.mouse.wheel(0, scroll_y)
                        await human_delay(1, 3)

                    # Extract HTML
                    html_content = await page.content()
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print(f"✅ HTML saved to {output_file}")

                    # parse and collect results (keeps your parsing logic)
                    price, model_number = parse_amazon_html(output_file)

                    await page.close()

                    results.append({"url": url, "file": output_file, "price": price, "model": model_number, "status": "ok"})
                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        await page.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"})

            # Save cookies/session state after all pages are processed
            storage_state = await context.storage_state()
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(storage_state, f, ensure_ascii=False, indent=4)
            print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

def parse_amazon_html(html_file_path="amazon.html"):
    # -------- Read HTML --------
    if not os.path.exists(html_file_path):
        print(f"Error: HTML file '{html_file_path}' not found.")
        return None, None

    with open(html_file_path, "r", encoding="utf-8") as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, "lxml")

    # -------- PRICE EXTRACTION (kept the same) --------
    price = None

    # Locate the element that holds the price information (the 'a-price-whole' class for the whole price)
    price_whole = soup.find("span", {"class": "a-price-whole"})  # The main price whole part
    price_fraction = soup.find("span", {"class": "a-price-fraction"})  # The decimal part of the price
    price_symbol = soup.find("span", {"class": "a-price-symbol"})  # The currency symbol

    # If we found the whole part and fraction part of the price
    if price_whole and price_fraction:
        # Safely get symbol text if present
        symbol_text = price_symbol.get_text().strip() if price_symbol else ""
        price = symbol_text + price_whole.get_text().strip() + "." + price_fraction.get_text().strip()

    # Print price or fallback message
    if price:
        print(f"The price of the product is: {price}")
    else:
        print("Price not found in the HTML file.")

    # -------- MODEL NUMBER EXTRACTION (kept the same) --------
    model_number = None

    # Find a <th> whose text contains "Item model number" (case-insensitive, trimmed)
    th_tag = soup.find(lambda tag: tag.name == "th" and "item model number" in tag.get_text(strip=True).lower())

    if th_tag:
        # Find the next <td> sibling that contains the model number
        td_tag = th_tag.find_next_sibling("td")
        if td_tag:
            model_number = td_tag.get_text(strip=True)

    # Print model number or fallback message
    if model_number:
        print(f"The model number is: {model_number}")
    else:
        print("Model number not found in the HTML file.")

    return price, model_number

# -----------------------
# BESTBUY-specific logic
# -----------------------
def parse_bestbuy_html(input_file="bestbuy.html"):
    # Load HTML file
    if not os.path.exists(input_file):
        print(f"Error: HTML file '{input_file}' not found.")
        return "Price not found", "Model number not found"

    with open(input_file, "r", encoding="utf-8") as f:
        html = f.read()

    soup = BeautifulSoup(html, "lxml")

    # -------- PRICE EXTRACTION --------
    price_element = soup.select_one('[data-testid="price-block-customer-price"] span')

    if price_element:
        price = price_element.get_text(strip=True)
    else:
        price = "Price not found"

    # -------- MODEL NUMBER EXTRACTION --------
    model_element = soup.select_one('.disclaimer .inline-block')

    if model_element:
        # model_element contains text like:  "Model: SM-S938UZBEXAA"
        model_number = model_element.get_text(strip=True).replace("Model:", "").strip()
    else:
        model_number = "Model number not found"

    # Print results
    print("\n--- Extracted Product Data (BestBuy) ---")
    print(f"File: {input_file}")
    print(f"Price: {price}")
    print(f"Model Number: {model_number}")
    print("--------------------------------\n")

    return price, model_number

async def save_bestbuy_htmls(
    urls,
    output_dir="outputs",
    cookies_file="bestbuy_cookies.json",
    headless=True,
):
    """
    Loop over list of BestBuy URLs, save each page's HTML to output_dir,
    parse with parse_bestbuy_html and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless, slow_mo=100)

        # Load existing cookies/session state if available
        if os.path.exists(cookies_file):
            print("🍪 Loading existing cookies/session...")
            context = await browser.new_context(storage_state=cookies_file)
        else:
            print("🆕 No cookies found, creating a new session...")
            context = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                viewport={"width": 1366, "height": 768},
            )

        results = []
        try:
            for idx, url in enumerate(urls, start=1):
                safe_name = sanitize_filename(url)[:120]
                output_file = os.path.join(output_dir, f"bestbuy_{idx}_{safe_name}.html")
                page = None
                try:
                    page = await context.new_page()
                    print(f"\n[BestBuy {idx}/{len(urls)}] Navigating to {url} ...")
                    await page.goto(url, wait_until="load")

                    # Wait randomly for page content to settle
                    await human_delay(3, 6)

                    # 🖱️ Simulate random human-like mouse movement
                    for _ in range(3):
                        x = random.randint(200, 800)
                        y = random.randint(200, 600)
                        await page.mouse.move(x, y, steps=random.randint(5, 15))
                        await human_delay(0.3, 1.5)

                    # 🖱️ Random scrolling
                    for _ in range(2):
                        scroll_y = random.randint(400, 1000)
                        await page.mouse.wheel(0, scroll_y)
                        await human_delay(1, 3)

                    # Extract HTML
                    html_content = await page.content()
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    print(f"✅ HTML saved to {output_file}")

                    # Parse and collect results (uses your exact parsing logic)
                    price, model = parse_bestbuy_html(output_file)

                    results.append({"url": url, "file": output_file, "price": price, "model": model, "status": "ok"})
                    await page.close()
                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        if page:
                            await page.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "model": None, "status": f"error: {e}"})

            # Save cookies/session state after processing all pages
            storage_state = await context.storage_state()
            with open(cookies_file, "w", encoding="utf-8") as f:
                json.dump(storage_state, f, ensure_ascii=False, indent=4)
            print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

# -----------------------
# SAMSUNG-specific logic
# -----------------------
async def wait_network_idle(page, timeout=15000):
    """Wait until network becomes idle (0 active requests)."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except TimeoutError:
        print("⚠️ networkidle timeout — continuing anyway")

def extract_sku_from_url(url: str):
    """Extract sku value from the given URL (looks for 'sku-<value>')."""
    if not url:
        return None
    m = re.search(r"sku-([A-Za-z0-9-]+)", url, re.IGNORECASE)
    if m:
        return m.group(1)
    return None

def extract_price(filename):
    """Extract 512GB model price from saved HTML using BeautifulSoup."""
    if not os.path.exists(filename):
        print(f"❌ File not found for parsing: {filename}")
        return None

    with open(filename, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    container = soup.find(id="device_info")
    if not container:
        print("❌ device_info not found in HTML")
        return None

    radios = container.find_all(attrs={"role": "radio"})
    target = None

    # Prefer aria-checked=true
    for r in radios:
        if r.get("aria-checked") == "true":
            target = r
            break

    # Otherwise find 512GB
    if target is None:
        for r in radios:
            if "512" in r.get_text():
                target = r
                break

    if not target:
        print("❌ Could not find 512GB radio")
        return None

    # Extract price
    text = target.get_text("\n", strip=True)
    prices = re.findall(r"\$\s*[\d,]+\.\d{2}", text)

    # Choose price that is NOT a "was:" value
    selected = None
    for line in text.split("\n"):
        if "$" in line and "was" not in line.lower():
            m = re.search(r"\$\s*[\d,]+\.\d{2}", line)
            if m:
                selected = m.group(0)
                break

    print("🔎 Extracted Price:", selected or (prices[-1] if prices else None))
    return selected or (prices[-1] if prices else None)

async def save_samsung_htmls(
    urls,
    output_dir="outputs",
    cookies_file="samsung_cookies.json",
    headless=True,
):
    """
    Loop over list of Samsung product URLs, save each page's HTML to output_dir,
    parse price and sku using the same logic you provided, and return results list.
    """
    os.makedirs(output_dir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)   # keep visible by default per original
        # Create or reuse context
        # if os.path.exists(cookies_file):
        #     print("🍪 Loading existing cookies/session...")
        #     context = await browser.new_context(storage_state=cookies_file)
        # else:
        print("🆕 No cookies found, creating a new session...")
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1600, "height": 900},
        )

        results = []
        try:
            for idx, url in enumerate(urls, start=1):
                safe_name = sanitize_filename(url)
                output_file = os.path.join(output_dir, f"samsung_{idx}_{safe_name}.html")

                page = None
                try:
                    page = await context.new_page()
                    print(f"\n[Samsung {idx}/{len(urls)}] Navigating to {url} ...")
                    await page.goto(url, wait_until="domcontentloaded")

                    print("Waiting for network to be idle...")
                    await wait_network_idle(page, timeout=20000)

                    print("Waiting for #device_info box...")
                    try:
                        await page.wait_for_selector("#device_info", timeout=20000)
                    except TimeoutError:
                        print("❌ #device_info did NOT load — Samsung blocked or loaded too slowly.")
                        # Still save HTML for debugging
                        html = await page.content()
                        with open(output_file, "w", encoding="utf-8") as f:
                            f.write(html)
                        sku = extract_sku_from_url(url)
                        results.append({"url": url, "file": output_file, "price": None, "sku": sku, "status": "partial: no device_info"})
                        await page.close()
                        continue

                    # Extra wait for prices inside #device_info
                    await page.wait_for_selector("#device_info span", timeout=15000)

                    # Save HTML
                    html = await page.content()
                    with open(output_file, "w", encoding="utf-8") as f:
                        f.write(html)
                    print(f"✅ HTML saved to {output_file}")

                    # Save cookies/session state after each page optionally (we'll write final at end too)
                    # storage = await context.storage_state()
                    # with open(cookies_file, "w", encoding="utf-8") as f:
                    #     json.dump(storage, f, indent=2)

                    # Parse saved HTML using your exact functions
                    price = extract_price(output_file)
                    sku = extract_sku_from_url(url)

                    print("🔎 Final extracted values — Price:", price, "SKU:", sku)
                    results.append({"url": url, "file": output_file, "price": price, "sku": sku, "status": "ok"})

                    # tiny cooperative yield
                    await human_delay_short()

                    await page.close()

                except Exception as e:
                    print(f"❌ Error processing URL {url}: {e}")
                    try:
                        if page:
                            await page.close()
                    except Exception:
                        pass
                    results.append({"url": url, "file": None, "price": None, "sku": None, "status": f"error: {e}"})

            # Write cookies/session state once more at the end
            # storage = await context.storage_state()
            # with open(cookies_file, "w", encoding="utf-8") as f:
            #     json.dump(storage, f, indent=2)
            # print(f"\n🍪 Cookies/session state written to {cookies_file}")

        finally:
            await browser.close()

    return results

# -----------------------
# Combined main
# -----------------------
async def main():
    # Replace/extend these lists with the product URLs you want to iterate over
    amazon_urls = [
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7JRKGH1/ref=sr_1_3?crid=2IIQ350CABWC7&keywords=galaxy%2Bz%2Bfold%2B7",
    "https://www.amazon.com/Smartphone-Unlocked-Processor-Manufacturer-Warranty/dp/B0DP3G4GVQ/ref=sr_1_1_sspa?crid=10TW4LFRAIOBO&dib=eyJ2IjoiMSJ9.uqQhueQzsbHe8zENbFmj7bUk0vIwEpi-0APakuwi3hHMu2vGmVltlmCoeqExLjwwHe1NY_y-eiRAZze4TELqwF9A5Z3q2WMC2EPG0p4nD5aGis4NWae_K-CRmvy0IwyOTABmJrdT_nBArRg_3HUXEeD8RiVcw9SrqiFQb-CKPztbZuf4z8k2ncgbVn8qKqGMwy7rSG9Br5vXcD_F-IobKCrdhThEoUQ0RDqrmYpPZPI.EY_aE-DTUSHzDMHcvx2u1pDmmaEgKrcYAQ31_D796hk&dib_tag=se&keywords=s25%2Bultra&qid=1763542293&sprefix=s25%2Caps%2C425&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1",
    "https://www.amazon.com/Samsung-Smartphone-Unlocked-Manufacturer-Warranty/dp/B0F7K3FZ79/ref=sr_1_1_sspa?crid=ZFFPIZBZ98GJ&dib=eyJ2IjoiMSJ9.UCywTCyyKG4bvq7perU6WJwDnwocjQBoU_CBTt0iLEilFUxs7eGFZYXZpU_ioObwnwWuyf6rjjxKURGHvrFykwP0YDyTNEHIJ6iMdK6L--UC4Xf9otHkBAGnuMrKXhDVPrKXBcX3EASPQMHPmIxeZyAUQDkEAC7kjvwYOc851BfCkl7yfIKNjFbb5-rq1n_ZNuEDFncqGGsmuRczfLMtzrq4HfNQyihnc2SfvotrbB8.gsaBvxdWlghc5_yM6ADh4JkaCZsCGwTYiwGF4rklYiw&dib_tag=se&keywords=galaxy%2Bz%2Bflip%2B7&qid=1763542634&sprefix=galaxy%2Bz%2Bflip7%2B%2Caps%2C427&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1",
    "https://www.amazon.com/SAMSUNG-Smartphone-Processor-ProScaler-Manufacturer/dp/B0DYVMVZSY/ref=sr_1_1_sspa?crid=18N7Z6JOQP2BV&dib=eyJ2IjoiMSJ9.nzLYfcsJ7KheFLAc8b9qkf36-GLK18wvNZAtoNJSu1Zuk0LTOxwvIqqD7blO0fqQDvGW1a_cFlDg5Nh6UJs25ksORtqvnynWCvMs3mnXvO49ZOy1Lc0OCa8xgu_zDwki3AucEZejB1tiHQzt8KYuAH3-YcGmTnO7s-Wn_1i_JPAcSstuLawUyqxRadquHocmToV-_PuNbtIeLyuTmsuGn88G4Hs_fJCfV7dzS_zI9l8.tjLmW8sTlnkCiJoCZPOQgL3qCTpiGvJ6gp9QKjm7R24&dib_tag=se&keywords=galaxy%2Bs25%2Bedge&qid=1763542733&sprefix=galaxy%2Bs25%2Bed%2Caps%2C351&sr=8-1-spons&sp_csd=d2lkZ2V0TmFtZT1zcF9hdGY&th=1"
    ]

    bestbuy_urls = [
        "https://www.bestbuy.com/product/samsung-galaxy-z-fold7-512gb-unlocked-blue-shadow/JJGRF3XK3P",
        "https://www.bestbuy.com/product/samsung-galaxy-s25-ultra-512gb-unlocked-titanium-black/J3ZYG25H6J",
        "https://www.bestbuy.com/product/samsung-galaxy-z-flip7-512gb-unlocked-jet-black/JJGRF335X6",
        "https://www.bestbuy.com/product/samsung-galaxy-s25-edge-512gb-unlocked-titanium-jet-black/JJGRF3CQKC"
    ]

    samsung_urls = [
    "https://www.samsung.com/us/smartphones/galaxy-z-fold7/buy/galaxy-z-fold7-512gb-unlocked-sku-sm-f966udbexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-ultra-512gb-unlocked-sku-sm-s938uzkexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-z-flip7/buy/galaxy-z-flip7-512gb-unlocked-sku-sm-f766uzkexaa/",
    "https://www.samsung.com/us/smartphones/galaxy-s25-ultra/buy/galaxy-s25-edge-512gb-unlocked-sku-sm-s937uzkexaa/"
    ]

    print("\n=== Running Amazon scraper ===")
    am_res = await save_amazon_htmls(amazon_urls, output_dir="outputs", cookies_file="amazon_cookies.json", headless=True)
    print("\nAmazon Summary:")
    for r in am_res:
        print(r)

    print("\n=== Running BestBuy scraper ===")
    bb_res = await save_bestbuy_htmls(bestbuy_urls, output_dir="outputs", cookies_file="bestbuy_cookies.json", headless=True)
    print("\nBestBuy Summary:")
    for r in bb_res:
        print(r)

    print("\n=== Running Samsung scraper ===")
    sam_res = await save_samsung_htmls(samsung_urls, output_dir="outputs", cookies_file="samsung_cookies.json", headless=True)
    print("\nSamsung Summary:")
    for r in sam_res:
        print(r)

    # -----------------------
    # Flatten results into a single dict: keys -> values
    # Keys format: "<site>_<index>_<keyname>" e.g. "amazon_1_url", "bestbuy_2_price"
    # -----------------------
    flat = {}
    flat["run_timestamp"] = datetime.utcnow().isoformat() + "Z"
    # Amazon
    for i, item in enumerate(am_res, start=1):
        for k, v in item.items():
            flat_key = f"amazon_{i}_{k}"
            flat[flat_key] = v
    # BestBuy
    for i, item in enumerate(bb_res, start=1):
        for k, v in item.items():
            flat_key = f"bestbuy_{i}_{k}"
            flat[flat_key] = v
    # Samsung
    for i, item in enumerate(sam_res, start=1):
        for k, v in item.items():
            flat_key = f"samsung_{i}_{k}"
            flat[flat_key] = v

    # Save flattened results to Excel (appends as new row)
    excel_file = os.path.join("outputs", "results.xlsx")
    save_dict_to_excel_row(flat, excel_file)

if __name__ == "__main__":
    asyncio.run(main())
    file_path = "outputs/results.xlsx"

    # column_references = [
    #     "a","d","ar","x","e","as","y","blank column",
    #     "i","aw","ac","j","ax","ad","blank column",
    #     "n","bb","ah","o","bc","ai","blank column",
    #     "s","bg","am","t","bh","an"
    # ]

    # created = copy_columns_by_references(
    #     file_path=file_path,
    #     column_refs=column_references,
    #     source_sheet_name=None,      # None => use first sheet; or set "Sheet1"
    #     new_sheet_base_name="SelectedColumns"
    # )
    # print(f"Created sheet: {created} in {file_path}")
